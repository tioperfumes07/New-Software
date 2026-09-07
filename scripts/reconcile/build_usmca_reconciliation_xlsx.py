#!/usr/bin/env python3
"""
Regenerate docs/reconcile/USMCA-RECONCILIATION-2026-09-06.xlsx from the data
measured on 2026-09-06 (Neon tiny-field-89581227 br-fancy-credit-akjnd07a,
bypass_rls=lucia, USMCA operating_company_id 5c854333-6ea5-4faa-af31-67cb272fef80)
and the owner source workbook IH35-BY-LOAD-20260904-WITH-DIESEL.xlsx
(sheets USMCA BY LOAD + DIESEL — LOAD NOT IN EXPORT).

Self-contained: the Neon-measured load table and the workbook settlement→load
map are embedded as literals so the workbook regenerates deterministically
without Neon or ~/Downloads. Provenance of every number is in the comments and
in the Methodology sheet. No network, no DB, no production writes.

Usage:  python3 scripts/reconcile/build_usmca_reconciliation_xlsx.py
Requires: openpyxl (pip install openpyxl)

LEAD RULING (2026-09-06 03:47Z) applied here, verbatim intent:
  - The void under IH35 LAW = status='cancelled' + cancel_reason, WORM-kept.
    soft_deleted_at is NOT the void mechanism. "Active" is counted by STATUS.
  - The 29 cancelled rows ARE the void: 21 pre-cutover + 8 Transportation-Faro
    (13509,13517,13524,13527,13531,13533,13539,13540) owner-decided 13:45Z.
  - NO re-quarantine, NO soft-delete. The earlier "78 active / KEEP-30" was a
    definition error (counted soft_deleted_at IS NULL instead of status).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "..", "..", "docs", "reconcile",
                   "USMCA-RECONCILIATION-2026-09-06.xlsx")

USMCA = "5c854333-6ea5-4faa-af31-67cb272fef80"
CUTOVER = "2026-08-07"

# ── Neon-measured load table (78 numeric USMCA loads), 2026-09-06 ──────────────
# cols: load, status, trip_type, driver, app_pickup (load_stops seq1), qbo_date
# (qbo_ar_invoices.txn_date where doc_number=load; mirror stops 2026-08-14 so
#  NULL after that is EXPECTED, not missing). "cls" = classification under the
#  lead ruling: VOID-TRANSP-PRECUTOVER | VOID-TRANSP-FARO | ACTIVE-USMCA | ASSIGNED-SB
LOADS = [
    ("13471","cancelled","NB","Neftali Coronado Urbano","2026-07-24","2026-07-24","VOID-TRANSP-PRECUTOVER"),
    ("13480","cancelled","TR","Neftali Coronado Urbano","2026-07-21","2026-07-28","VOID-TRANSP-PRECUTOVER"),
    ("13482","cancelled","TR","Leonel Antonio Morales Noguez","2026-07-28","2026-07-31","VOID-TRANSP-PRECUTOVER"),
    ("13484","cancelled","NB","Ruben Pedro Perez Garcia","2026-07-31","2026-07-31","VOID-TRANSP-PRECUTOVER"),
    ("13485","cancelled","TR","Leonel Antonio Morales Noguez","2026-07-31","2026-08-03","VOID-TRANSP-PRECUTOVER"),
    ("13486","cancelled","NB","Jose Miguel de Santiago Palacios","2026-07-31","2026-08-03","VOID-TRANSP-PRECUTOVER"),
    ("13487","cancelled","TR","Jorge Luis Infante Corona","2026-07-28","2026-07-31","VOID-TRANSP-PRECUTOVER"),
    ("13488","cancelled","TR","Jose Miguel de Santiago Palacios","2026-07-28","2026-07-30","VOID-TRANSP-PRECUTOVER"),
    ("13491","cancelled","TR","Ruben Pedro Perez Garcia","2026-07-28","2026-07-31","VOID-TRANSP-PRECUTOVER"),
    ("13492","cancelled","TR","Neftali Coronado Urbano","2026-07-31","2026-08-03","VOID-TRANSP-PRECUTOVER"),
    ("13493","cancelled","TR","Jorge Luis Infante Corona","2026-07-31","2026-08-03","VOID-TRANSP-PRECUTOVER"),
    ("13494","cancelled","TR","Hugo Gaytan","2026-07-31","2026-08-03","VOID-TRANSP-PRECUTOVER"),
    ("13495","cancelled","TR","Jose Antonio Vicente Martinez","2026-07-31","2026-08-03","VOID-TRANSP-PRECUTOVER"),
    ("13496","cancelled","TR","Jose Antonio Vicente Martinez","2026-08-03","2026-08-04","VOID-TRANSP-PRECUTOVER"),
    ("13497","cancelled","NB","Concepcion Cordova Dominguez","2026-07-03","2026-08-03","VOID-TRANSP-PRECUTOVER"),  # app 07/03 = AlwaysTrack error; QBO 08/03 governs
    ("13498","cancelled","TR","Angel Alfonso Sosa","2026-08-03","","VOID-TRANSP-PRECUTOVER"),
    ("13499","cancelled","TR","Neftali Coronado Urbano","2026-07-21","2026-08-04","VOID-TRANSP-PRECUTOVER"),      # app 07/21 = AlwaysTrack error; QBO 08/04 governs
    ("13500","cancelled","TR","Hugo Gaytan","2026-08-03","2026-08-04","VOID-TRANSP-PRECUTOVER"),
    ("13503","cancelled","TR","Neftali Coronado Urbano","2026-08-04","2026-08-07","VOID-TRANSP-PRECUTOVER"),      # QBO 08/07 noted; owner ruled pre-cutover 13:36Z
    ("13504","cancelled","TR","Jorge Luis Infante Corona","2026-08-04","2026-08-07","VOID-TRANSP-PRECUTOVER"),   # QBO 08/07 noted; owner ruled pre-cutover 13:36Z
    ("13506","cancelled","NB","Alfonso Hidalgo Chavez","2026-08-04","2026-08-07","VOID-TRANSP-PRECUTOVER"),      # QBO 08/07 noted; owner ruled pre-cutover 13:36Z
    ("13508","assigned_not_dispatched","SB","Angel Alfonso Sosa","2026-08-07","2026-08-07","ASSIGNED-SB"),
    ("13509","cancelled","TR","Neftali Coronado Urbano","2026-08-07","2026-08-10","VOID-TRANSP-FARO"),           # owner-decided 13:45Z
    ("13510","dispatched","TR","Jorge Luis Infante Corona","2026-08-07","2026-08-10","ACTIVE-USMCA"),
    ("13511","dispatched","TR","Concepcion Cordova Dominguez","2026-08-07","2026-08-10","ACTIVE-USMCA"),
    ("13512","dispatched","NB","Pedro Abraham Lopez Collado","2026-08-10","","ACTIVE-USMCA"),
    ("13513","dispatched","TR","Pedro Abraham Lopez Collado","2026-08-12","","ACTIVE-USMCA"),
    ("13514","dispatched","TR","Alfonso Hidalgo Chavez","2026-08-10","","ACTIVE-USMCA"),
    ("13515","dispatched","TR","Leonel Antonio Morales Noguez","2026-08-13","","ACTIVE-USMCA"),
    ("13516","dispatched","TR","Alfonso Hidalgo Chavez","2026-08-13","","ACTIVE-USMCA"),
    ("13517","cancelled","NB","Jose Antonio Vicente Martinez","2026-08-07","","VOID-TRANSP-FARO"),               # owner-decided 13:45Z
    ("13518","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-11","","ACTIVE-USMCA"),
    ("13519","dispatched","NB","Jorge Luis Infante Corona","2026-08-11","","ACTIVE-USMCA"),
    ("13520","dispatched","TR","Leonel Antonio Morales Noguez","2026-08-11","","ACTIVE-USMCA"),
    ("13521","dispatched","TR","Jorge Luis Infante Corona","2026-08-16","","ACTIVE-USMCA"),
    ("13522","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-15","","ACTIVE-USMCA"),
    ("13523","dispatched","NB","Leonel Antonio Morales Noguez","2026-08-15","","ACTIVE-USMCA"),
    ("13524","cancelled","TR","Hugo Gaytan","2026-08-14","","VOID-TRANSP-FARO"),                                 # owner-decided 13:45Z
    ("13525","dispatched","TR","Hugo Gaytan","2026-08-07","","ACTIVE-USMCA"),
    ("13526","dispatched","NB","Luis Armando Sosa Perez","2026-08-18","","ACTIVE-USMCA"),
    ("13527","cancelled","TR","Luis Armando Sosa Perez","2026-08-14","","VOID-TRANSP-FARO"),                     # owner-decided 13:45Z
    ("13528","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-18","","ACTIVE-USMCA"),
    ("13529","dispatched","NB","Hugo Gaytan","2026-08-17","","ACTIVE-USMCA"),
    ("13530","dispatched","NB","Rafael Rogelio Rivero Reynoso","2026-08-18","","ACTIVE-USMCA"),
    ("13531","cancelled","NB","Genaro Guerrero Chavez","2026-08-17","","VOID-TRANSP-FARO"),                      # owner-decided 13:45Z
    ("13532","dispatched","TR","Rafael Rogelio Rivero Reynoso","2026-08-20","","ACTIVE-USMCA"),
    ("13533","cancelled","TR","Concepcion Cordova Dominguez","2026-08-19","","VOID-TRANSP-FARO"),                # owner-decided 13:45Z
    ("13534","dispatched","TR","Leonel Antonio Morales Noguez","2026-08-19","","ACTIVE-USMCA"),
    ("13535","dispatched","TR","Jorge Luis Infante Corona","2026-08-18","","ACTIVE-USMCA"),
    ("13536","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-20","","ACTIVE-USMCA"),
    ("13537","dispatched","TR","Jorge Luis Infante Corona","2026-08-21","","ACTIVE-USMCA"),
    ("13538","dispatched","TR","Genaro Guerrero Chavez","2026-08-21","","ACTIVE-USMCA"),
    ("13539","cancelled","NB","Angel Alfonso Sosa","2026-08-20","","VOID-TRANSP-FARO"),                          # owner-decided 13:45Z
    ("13540","cancelled","TR","Hugo Gaytan","2026-08-22","","VOID-TRANSP-FARO"),                                 # owner-decided 13:45Z
    ("13541","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-25","","ACTIVE-USMCA"),
    ("13542","dispatched","TR","Leonel Antonio Morales Noguez","2026-08-25","","ACTIVE-USMCA"),
    ("13543","dispatched","TR","Genaro Guerrero Chavez","2026-08-24","","ACTIVE-USMCA"),
    ("13544","dispatched","TR","Rafael Rogelio Rivero Reynoso","2026-08-24","","ACTIVE-USMCA"),
    ("13545","dispatched","TR","Hugo Gaytan","2026-08-25","","ACTIVE-USMCA"),
    ("13546","dispatched","TR","Angel Alfonso Sosa","2026-08-26","","ACTIVE-USMCA"),
    ("13547","dispatched","TR","Genaro Guerrero Chavez","2026-08-26","","ACTIVE-USMCA"),
    ("13548","dispatched","TR","Concepcion Cordova Dominguez","2026-08-24","","ACTIVE-USMCA"),
    ("13549","dispatched","TR","Alfonso Hidalgo Chavez","2026-08-25","","ACTIVE-USMCA"),
    ("13550","dispatched","TR","Jorge Luis Infante Corona","2026-08-26","","ACTIVE-USMCA"),
    ("13551","dispatched","NB","Vicente Santos Contreras","2026-08-26","","ACTIVE-USMCA"),
    ("13552","dispatched","TR","Angel Alfonso Sosa","2026-08-28","","ACTIVE-USMCA"),
    ("13554","dispatched","TR","Leonel Antonio Morales Noguez","2026-08-28","","ACTIVE-USMCA"),
    ("13555","dispatched","TR","Alfonso Hidalgo Chavez","2026-08-19","","ACTIVE-USMCA"),
    ("13557","dispatched","TR","Jorge Luis Infante Corona","2026-08-28","","ACTIVE-USMCA"),
    ("13558","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-28","","ACTIVE-USMCA"),
    ("13559","dispatched","TR","Genaro Guerrero Chavez","2026-08-29","","ACTIVE-USMCA"),
    ("13560","dispatched","TR","Hugo Gaytan","2026-08-29","","ACTIVE-USMCA"),
    ("13561","dispatched","TR","Luis Armando Sosa Perez","2026-08-29","","ACTIVE-USMCA"),
    ("13562","dispatched","TR","Genaro Guerrero Chavez","2026-09-01","","ACTIVE-USMCA"),
    ("13565","dispatched","TR","Neftali Coronado Urbano","2026-08-25","","ACTIVE-USMCA"),
    ("13566","dispatched","TR","Neftali Coronado Urbano","2026-08-28","","ACTIVE-USMCA"),
    ("13567","dispatched","TR","Luis Armando Sosa Perez","2026-08-28","","ACTIVE-USMCA"),
    ("13568","dispatched","TR","Jose Antonio Vicente Martinez","2026-08-31","","ACTIVE-USMCA"),
]

# ── Settlement → load map, extracted from IH35-BY-LOAD-20260904-WITH-DIESEL.xlsx
#    (USMCA BY LOAD sheet LOAD rows; 5791/5792 from the DIESEL sheet). 09-04 snapshot. ──
SETTLEMENT_MAP = [
    ("5769", ["13508"]),
    ("5771", ["13510"]),
    ("5772", ["13512","13513"]),
    ("5773", ["13511"]),
    ("5774", ["13518"]),
    ("5775", ["13514","13516"]),
    ("5776", ["13520"]),
    ("5777", ["13519","13521"]),
    ("5779", ["13526"]),
    ("5780", ["13532"]),
    ("5781", ["13523","13534"]),
    ("5782", ["13529"]),
    ("5783", ["13535","13537"]),
    ("5784", ["13528","13536"]),
    ("5785", ["13538","13543"]),
    ("5786", ["13548"]),
    ("5787", ["13549"]),
    ("5791", ["13560"]),   # DIESEL sheet
    ("5792", ["13559"]),   # DIESEL sheet
    ("(unassigned in 09-04 snapshot)", ["13541","13542","13544","13545","13546","13547","13550","13551","13552","13554","13556","13557"]),
]

# ── Owner manual-entry HOLD (lead map; 5766 = Transportation, excluded) ──
OWNER_HOLD = [
    ("5772","13512"),("5772","13513"),("5776","13520"),("5780","13532"),
    ("5783","13535"),("5783","13537"),("5784","13528"),("5784","13536"),
]

# ── Missing / not-in-app loads (lead ruling) ──
MISSING = [
    ("13556","USMCA BY LOAD (Hummingbird Logistix, Laredo→Medley FL, no settlement #, no date, T176)","TO SEED — find signed source (settlement PDF / Faro / QBO invoice), hand CC-3 the seed row"),
    ("13553","TRANSPORTATION BY LOAD (PAYPA Transport)","DO NOT SEED — correctly Transportation"),
    ("13505","TRANSPORTATION BY LOAD (lead 2026-09-06)","DO NOT SEED — Transportation; OWNER hand settlement 5776 (pickup 08-03)"),
    ("13507","TRANSPORTATION BY LOAD (lead 2026-09-06)","DO NOT SEED — Transportation; OWNER hand settlement 5772 (pickup 08-06)"),
    ("13563","neither BY-LOAD sheet","Cursor saw these ONLY as numeric-sequence gaps between 13562 and 13565 in the Neon USMCA set; NOT on any workbook sheet"),
    ("13564","neither BY-LOAD sheet","Cursor saw these ONLY as numeric-sequence gaps between 13562 and 13565 in the Neon USMCA set; NOT on any workbook sheet"),
]

# ── styling ──
H = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="14314F")
SUB = Font(bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*(Side(style="thin", color="E5E7EB"),)*4)
CLS_FILL = {
    "VOID-TRANSP-PRECUTOVER": PatternFill("solid", fgColor="FDE8E8"),
    "VOID-TRANSP-FARO":       PatternFill("solid", fgColor="FEF3C7"),
    "ACTIVE-USMCA":           PatternFill("solid", fgColor="E7F5EC"),
    "ASSIGNED-SB":            PatternFill("solid", fgColor="E5EDF5"),
}

def hdr(ws, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i, c); cell.font = H; cell.fill = HFILL; cell.alignment = WRAP; cell.border = THIN

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# 1) SUMMARY
ws = wb.active; ws.title = "Summary"
disp = sum(1 for l in LOADS if l[1]=="dispatched")
canc = sum(1 for l in LOADS if l[1]=="cancelled")
asg  = sum(1 for l in LOADS if l[1]=="assigned_not_dispatched")
pre  = sum(1 for l in LOADS if l[6]=="VOID-TRANSP-PRECUTOVER")
faro = sum(1 for l in LOADS if l[6]=="VOID-TRANSP-FARO")
rows = [
    ["IH35 USMCA ↔ Transportation reconciliation", "2026-09-06 (Cursor, registrar)"],
    ["Neon", "tiny-field-89581227 / br-fancy-credit-akjnd07a / bypass_rls=lucia"],
    ["USMCA operating_company_id", USMCA],
    ["Cutover floor", "pickup ≥ %s = USMCA" % CUTOVER],
    ["", ""],
    ["Numeric USMCA loads (total)", len(LOADS)],
    ["  dispatched (ACTIVE USMCA)", disp],
    ["  cancelled = VOID (WORM, status+reason, NOT soft-delete)", canc],
    ["    of which pre-cutover Transportation", pre],
    ["    of which Transportation-Faro (owner-decided 13:45Z)", faro],
    ["  assigned_not_dispatched (SB 13508)", asg],
    ["", ""],
    ["LEAD RULING 2026-09-06", "cancelled+reason IS the void; 'active' counted by STATUS not soft_deleted_at; NO re-quarantine, NO soft-delete; KEEP-30 list retired."],
    ["Prior Cursor error", "'78 active' counted soft_deleted_at IS NULL — a definition error. Corrected."],
    ["Settlement numbers", "5769–5795 are PAPER/Excel tour numbers (workbook), not DB IDs (DB = S-13642…S-13656, unlinked)."],
    ["SB legs", "exactly 1 SB in USMCA (13508); no SB returns seeded — confirm from Excel before seeding."],
    ["QBO", "doc_number = load number; txn_date governs the AlwaysTrack gap; mirror stale after 2026-08-14 (single Transportation realm 91e0bf0a)."],
]
for r in rows:
    ws.append(r)
ws["A1"].font = Font(bold=True, size=13)
for c in range(1, 3):
    ws.cell(6, 1).font = SUB
autosize(ws, [56, 66])

# 2) METHODOLOGY / LOGIC
ws = wb.create_sheet("Logic")
logic = [
    ["#","Rule","Detail"],
    ["1","Entity = pickup date vs cutover","pickup ≥ 2026-08-07 → USMCA; earlier → Transportation. Owner floor."],
    ["2","AlwaysTrack gap → QuickBooks corroborates","AlwaysTrack (Samsara) was down a few days end-Jul/early-Aug, so app pickup dates in that window are unreliable. QBO doc_number = load number and helps recover a load's identity. NOTE (lead 2026-09-06): QBO txn_date is the INVOICE date, NOT the pickup — the entity rule keys on PICKUP ≥ 08/07 and 'not Transportation-Faro'. QBO 08/07 on 13503/04/06 is the invoice date; their pickup is 08-04 → Transportation."],
    ["3","Void = status, not delete","IH35 LAW 'void, never delete'. The void = status='cancelled' + cancel_reason ('WRONG ENTITY — TRANSPORTATION …' / Transportation-Faro), canceled_at set, row RETAINED (soft_deleted_at NULL). WORM = the cancel register."],
    ["4","'Active' counted by STATUS","Active USMCA = status='dispatched' (+ the one 'assigned_not_dispatched' SB). Never 'soft_deleted_at IS NULL' — that was the prior definition error that produced '78 active'."],
    ["5","Owner-decided Faro void 13:45Z","13509,13517,13524,13527,13531,13533,13539,13540 were voided that afternoon as Transportation-Faro. They are NOT mislabelled errors; they are owner decisions. Not re-activated."],
    ["6","Settlement↔load map = the workbook","Source = IH35-BY-LOAD-20260904-WITH-DIESEL.xlsx (USMCA BY LOAD + DIESEL sheets). DB settlements carry no 5769–5795; those are paper tour numbers."],
    ["7","No production writes","This reconciliation is documentation only. Any reclassify/seed is a separate owner-approved action through the service layer."],
    ["8","SB returns not invented","Only 13508 is SB. The NB+TR tours have no seeded SB closing leg. Confirm from the signed Excel before any SB seed; never fabricate."],
]
for r in logic: ws.append(r)
hdr(ws, logic[0])
autosize(ws, [4, 34, 96])
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = WRAP; c.border = THIN

# 3) ALL LOADS (Neon-measured)
ws = wb.create_sheet("All USMCA Loads")
cols = ["Load #","Status","Trip","Driver","App pickup (load_stops)","QBO txn_date","Classification"]
hdr(ws, cols)
for l in sorted(LOADS, key=lambda x:int(x[0])):
    ws.append([l[0], l[1], l[2], l[3], l[4], l[5] or "—", l[6]])
    fill = CLS_FILL.get(l[6])
    if fill:
        for cix in range(1, len(cols)+1):
            ws.cell(ws.max_row, cix).fill = fill
autosize(ws, [10, 22, 6, 30, 22, 14, 24])

# 4) VOID — Transportation (the 29)
ws = wb.create_sheet("Void (Transportation)")
hdr(ws, ["Load #","Bucket","Trip","Driver","App pickup","QBO date","Note"])
for l in sorted([x for x in LOADS if x[1]=="cancelled"], key=lambda x:int(x[0])):
    note = ""
    if l[0] in ("13503","13504","13506"): note = "QBO 08/07 (informational); owner ruled pre-cutover 13:36Z — kept as void"
    if l[0] in ("13497","13499"): note = "app date is an AlwaysTrack-gap error; QBO date governs"
    if l[6]=="VOID-TRANSP-FARO": note = "Transportation-Faro, owner-decided 13:45Z"
    ws.append([l[0], "pre-cutover" if l[6]=="VOID-TRANSP-PRECUTOVER" else "Transportation-Faro", l[2], l[3], l[4], l[5] or "—", note])
autosize(ws, [10, 18, 6, 30, 14, 12, 52])
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = WRAP; c.border = THIN

# 5) ACTIVE USMCA (dispatched + SB)
ws = wb.create_sheet("Active USMCA")
hdr(ws, ["Load #","Status","Trip","Driver","Pickup"])
for l in sorted([x for x in LOADS if x[6] in ("ACTIVE-USMCA","ASSIGNED-SB")], key=lambda x:int(x[0])):
    ws.append([l[0], l[1], l[2], l[3], l[4]])
autosize(ws, [10, 24, 6, 30, 14])

# 6) SETTLEMENT → LOAD map
ws = wb.create_sheet("Settlement Map")
hdr(ws, ["Settlement # (paper)","Loads","Source"])
for s, lds in SETTLEMENT_MAP:
    ws.append([s, ", ".join(lds), "USMCA BY LOAD" if s[:1]=="5" else "workbook (no settl# in 09-04 snapshot)"])
autosize(ws, [30, 40, 40])
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = WRAP; c.border = THIN

# 7) OWNER HOLD
ws = wb.create_sheet("Owner HOLD")
hdr(ws, ["Settlement #","Load #","Status in Neon"])
by = {l[0]: l for l in LOADS}
for s, ld in OWNER_HOLD:
    ws.append([s, ld, by.get(ld, ("","(not found)"))[1]])
autosize(ws, [16, 12, 24])

# 8) MISSING / TO-SEED
ws = wb.create_sheet("Missing")
hdr(ws, ["Load #","Where seen","Action"])
for r in MISSING: ws.append(list(r))
autosize(ws, [10, 44, 70])
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = WRAP; c.border = THIN

# 9) QBO GAP CROSS-CHECK
ws = wb.create_sheet("QBO Gap Cross-check")
hdr(ws, ["Load #","App pickup","QBO txn_date","Reading"])
gap = [l for l in LOADS if l[5] and l[4] != l[5]]
for l in sorted(gap, key=lambda x:int(x[0])):
    rd = "QBO INVOICE date later than app pickup — AlwaysTrack gap distortion"
    if l[5]=="2026-08-07": rd = "QBO INVOICE date 08/07 (NOT pickup); pickup 08-04 → Transportation (lead 2026-09-06)"
    ws.append([l[0], l[4], l[5], rd])
autosize(ws, [10, 14, 14, 52])
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = WRAP; c.border = THIN

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("wrote", os.path.relpath(OUT))
print("sheets:", wb.sheetnames)
print(f"totals: dispatched={disp} cancelled={canc} (pre-cutover={pre} faro={faro}) assigned_sb={asg} total={len(LOADS)}")
